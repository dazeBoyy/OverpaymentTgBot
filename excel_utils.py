from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

async def generate_excel(data, accounting_name):
    # Создаем новый Excel файл
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Учёт данных"

    # Заголовки
    headers = ["Дата", "Выполняемая работа", "Сумма", "Бюджет"]
    worksheet.append(headers)

    # Форматирование заголовков
    header_fill = PatternFill(start_color="FEF1E0", end_color="FEF1E0", fill_type="solid")
    header_font = Font(name="Calibri", size=11, color="FF0000", bold=True, italic=True)  # Жирный и курсив
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for cell in worksheet[1]:  # Проходим по заголовкам (первой строке)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание заголовков по центру
        cell.border = thin_border  # Добавляем границы к заголовкам

    # Настройка стиля границы
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Добавляем данные с применением стилей
    data_font = Font(name="Arial")
    for entry in data:
        row = [entry['date'], entry['work'], entry['sum'], entry['budget']]
        worksheet.append(row)

        # Применяем шрифт Arial и границы ко всем ячейкам данных
        for col in worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = data_font  # Применяем шрифт к каждой ячейке
                cell.border = thin_border  # Применяем границу к каждой ячейке
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Итоговая строка
    last_row = len(data) + 6  # +2 из-за заголовков
    worksheet[f"C{last_row}"] = "Сумма:"  # Заголовок для итоговой строки
    worksheet[f"D{last_row}"] = f'=SUM(C2:C{last_row-1})'  # Сумма по столбцу "Сумма"
    worksheet[f"C{last_row}"].border = thin_border
    worksheet[f"C{last_row}"].alignment = Alignment(horizontal='center',
                                                    vertical='center')  # Выравнивание суммы по центру
    worksheet[f"D{last_row}"].border = thin_border
    worksheet[f"D{last_row}"].alignment = Alignment(horizontal='center',
                                                    vertical='center')  # Выравнивание итоговой суммы по центру

    # Применяем границы к итоговой строке
    for col in worksheet.iter_cols(min_row=last_row, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.border = thin_border  # Применяем границу к итоговой ячейке

    # Применяем границы ко всем ячейкам между данными и итоговой суммой
    for row in range(2, last_row):  # Проходим по всем строкам с данными
        for col in worksheet.iter_cols(min_row=row, max_row=row, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.border = thin_border  # Применяем границу к каждой ячейке

    # Опционально: Задайте ширину столбцов
    column_widths = [17.29, 54.63, 26.75, 41.13]  # Ширина столбцов для заголовков
    for i, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[chr(64 + i)].width = width  # chr(64 + i) превращает 1 в 'A', 2 в 'B' и так далее

    # Сохранение файла
    excel_filename = f"{accounting_name}.xlsx"
    workbook.save(excel_filename)

    return os.path.abspath(excel_filename)  # Возвращаем полный путь к файлу
